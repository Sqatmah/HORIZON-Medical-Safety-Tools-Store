from io import BytesIO
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import permissions
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .reports import (
    IsAdminOnly, SalesReportView, ProductsReportView,
    CustomersReportView, OrdersReportView,
)


HEADER_FILL = PatternFill(start_color="1A4B4C", end_color="1A4B4C", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(color="1A4B4C", bold=True, size=16)
THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'),
)


def style_sheet(ws, title, headers, rows):
    ws.sheet_view.rightToLeft = True

    ws.merge_cells('A1:' + get_column_letter(len(headers)) + '1')
    ws['A1'] = 'مؤسسة الابتكار التقني - Tech Innovation'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:' + get_column_letter(len(headers)) + '2')
    ws['A2'] = title
    ws['A2'].font = Font(bold=True, size=13, color="00A8CC")
    ws['A2'].alignment = Alignment(horizontal='center')

    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


class ExportExcelView(APIView):
    permission_classes = [IsAdminOnly]

    def get(self, request, report_type):
        wb = Workbook()
        ws = wb.active

        if report_type == 'sales':
            data = SalesReportView().get(request).data
            ws.title = 'تقرير المبيعات'
            headers = ['المؤشر', 'القيمة']
            rows = [
                ['عدد الطلبات', data['total_orders']],
                ['إجمالي الإيرادات (ريال)', float(data['total_revenue'])],
                ['إجمالي الضريبة المحصلة (ريال)', float(data['total_vat_collected'])],
                ['متوسط قيمة الطلب (ريال)', round(float(data['average_order_value']), 2)],
            ]
            style_sheet(ws, f"تقرير المبيعات - آخر {data['period_days']} يوم", headers, rows)
            filename = 'sales_report.xlsx'

        elif report_type == 'products':
            data = ProductsReportView().get(request).data
            ws.title = 'تقرير المنتجات'
            headers = ['المنتج', 'SKU', 'السعر', 'المخزون', 'عدد المبيعات', 'التقييم', 'الحالة', 'التصنيف']
            rows = [[p['name_ar'], p['sku'], float(p['price']), p['stock'], p['quantity_sold'], p['rating_avg'], p['status'], p['category']] for p in data]
            style_sheet(ws, 'تقرير المنتجات التفصيلي', headers, rows)
            filename = 'products_report.xlsx'

        elif report_type == 'customers':
            data = CustomersReportView().get(request).data
            ws.title = 'تقرير العملاء'
            headers = ['اسم المستخدم', 'البريد الإلكتروني', 'الهاتف', 'نوع العميل', 'عدد الطلبات', 'إجمالي الإنفاق']
            rows = [[c['username'], c['email'], c['phone'], c['customer_type'], c['total_orders'], float(c['total_spent'])] for c in data]
            style_sheet(ws, 'تقرير العملاء', headers, rows)
            filename = 'customers_report.xlsx'

        elif report_type == 'orders':
            data = OrdersReportView().get(request).data
            ws.title = 'تقرير الطلبات'
            headers = ['رقم الطلب', 'البريد الإلكتروني', 'الحالة', 'حالة الدفع', 'طريقة الدفع', 'المجموع']
            rows = [[o['order_number'], o['customer_email'], o['status'], o['payment_status'], o['payment_method'], float(o['total'])] for o in data]
            style_sheet(ws, 'تقرير الطلبات التفصيلي', headers, rows)
            filename = 'orders_report.xlsx'

        else:
            return HttpResponse('نوع تقرير غير معروف', status=400)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response